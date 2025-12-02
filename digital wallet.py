import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os



balance = 0
transactions = []
correct_pin = "1234"
file_name = "wallet_transactions.xlsx"



def load_transactions_from_excel():
    global transactions, balance

    if not os.path.exists(file_name):
       
        wb = Workbook()
        ws = wb.active
        ws.append(["Type", "Receiver", "Amount"])
        wb.save(file_name)
        return

    wb = load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        tran_type, receiver, amt = row

        if tran_type == "ADD":
            transactions.append(("ADD", amt))
            balance += amt

        elif tran_type == "PAY":
            transactions.append(("PAY", receiver, amt))
            balance -= amt


def save_transaction_to_excel(tran_type, receiver, amount):
    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([tran_type, receiver, amount])
    wb.save(file_name)


def login_window():
    login = tk.Tk()
    login.title("Digital Wallet - Login")
    login.geometry("300x200")

    tk.Label(login, text="Enter PIN", font=("Arial", 14)).pack(pady=10)

    pin_entry = tk.Entry(login, show="*", font=("Arial", 14))
    pin_entry.pack(pady=5)

    def check_pin():
        if pin_entry.get() == correct_pin:
            login.destroy()
            main_window()
        else:
            messagebox.showerror("Error", "Incorrect PIN!")

    tk.Button(login, text="Login", command=check_pin, font=("Arial", 12)).pack(pady=15)

    login.mainloop()


def main_window():
    win = tk.Tk()
    win.title("Digital Wallet - FinTech Project")
    win.geometry("400x350")

    tk.Label(win, text="DIGITAL WALLET", font=("Arial", 18, "bold")).pack(pady=10)

   
    def check_balance():
        messagebox.showinfo("Balance", f"Current Balance: {balance} PKR")


    def add_money():
        add_win = tk.Toplevel(win)
        add_win.title("Add Money")
        add_win.geometry("300x200")

        tk.Label(add_win, text="Enter Amount:", font=("Arial", 12)).pack(pady=10)
        amount_entry = tk.Entry(add_win, font=("Arial", 12))
        amount_entry.pack()

        def add_amount():
            global balance
            amount = int(amount_entry.get())
            balance += amount

            transactions.append(("ADD", amount))

           
            save_transaction_to_excel("ADD", "", amount)

            messagebox.showinfo("Success", "Amount Added!")
            add_win.destroy()

        tk.Button(add_win, text="Add", command=add_amount).pack(pady=20)

 
    def pay_money():
        pay_win = tk.Toplevel(win)
        pay_win.title("Pay Money")
        pay_win.geometry("300x250")

        tk.Label(pay_win, text="Receiver Name:", font=("Arial", 12)).pack(pady=5)
        name_entry = tk.Entry(pay_win, font=("Arial", 12))
        name_entry.pack()

        tk.Label(pay_win, text="Amount:", font=("Arial", 12)).pack(pady=5)
        amount_entry = tk.Entry(pay_win, font=("Arial", 12))
        amount_entry.pack()

        def send_payment():
            global balance
            receiver = name_entry.get()
            amount = int(amount_entry.get())

            if amount > balance:
                messagebox.showerror("Error", "Not enough balance!")
            else:
                balance -= amount

                transactions.append(("PAY", receiver, amount))

                save_transaction_to_excel("PAY", receiver, amount)

                messagebox.showinfo("Success", f"Paid {amount} PKR to {receiver}!")
                pay_win.destroy()

        tk.Button(pay_win, text="Send", command=send_payment).pack(pady=20)

    def show_history():
        history_win = tk.Toplevel(win)
        history_win.title("Transaction History")
        history_win.geometry("350x300")

        tk.Label(history_win, text="Transaction History", font=("Arial", 14, "bold")).pack(pady=10)

        if not transactions:
            tk.Label(history_win, text="No transactions yet.", font=("Arial", 12)).pack()
            return

        for i, t in enumerate(transactions, start=1):
            if t[0] == "ADD":
                text = f"{i}. Added {t[1]} PKR"
            else:
                text = f"{i}. Paid {t[2]} PKR to {t[1]}"

            tk.Label(history_win, text=text, font=("Arial", 11)).pack(anchor="w", padx=20)
    def clear_all_data():
        global balance, transactions

        confirm = messagebox.askyesno("Confirm", "Are you sure you want to CLEAR all data?")
        if not confirm:
            return

        balance = 0
        transactions = []

        wb = Workbook()
        ws = wb.active
        ws.append(["Type", "Receiver", "Amount"])
        wb.save(file_name)

        messagebox.showinfo("Cleared", "All data has been cleared!")

   
    tk.Button(win, text="Check Balance", width=20, command=check_balance).pack(pady=5)
    tk.Button(win, text="Add Money", width=20, command=add_money).pack(pady=5)
    tk.Button(win, text="Pay Money", width=20, command=pay_money).pack(pady=5)
    tk.Button(win, text="Transaction History", width=20, command=show_history).pack(pady=5)
    tk.Button(win, text="Clear All Data", width=20, command=clear_all_data).pack(pady=5)
    tk.Button(win, text="Exit", width=20, command=win.destroy).pack(pady=15)
   


    win.mainloop()


load_transactions_from_excel()
login_window()
