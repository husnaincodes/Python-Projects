from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── COLOURS ──────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_BLUE    = "1B4F72"
ACCENT_BLUE = "2E86C1"
LIGHT_BLUE  = "D6EAF8"
ACCENT_GOLD = "F39C12"
ACCENT_GREEN= "1E8449"
ACCENT_RED  = "C0392B"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F3F4"
MED_GRAY    = "BDC3C7"
DARK_GRAY   = "7F8C8D"
BLACK       = "000000"
INPUT_BLUE  = "0000FF"
LINK_GREEN  = "008000"

def thin():
    s = Side(style='thin', color=MED_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def bold_border():
    s = Side(style='medium', color=BLACK)
    return Border(bottom=s)

def hfill(color): return PatternFill('solid', fgColor=color)

def hdr_row(ws, row, cols, bg=MID_BLUE, h=20):
    for c, t in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
        cell.fill = hfill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin()
    ws.row_dimensions[row].height = h

def title(ws, row, sc, ec, text, bg=DARK_NAVY, sz=13, h=30, fc=WHITE):
    ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
    c = ws.cell(row=row, column=sc, value=text)
    c.font = Font(bold=True, size=sz, color=fc, name='Calibri')
    c.fill = hfill(bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = h

def wcols(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

def fmt_cell(ws, r, c, val=None, bold=False, color=BLACK, bg=None, align='left',
             num_fmt=None, sz=10, italic=False, wrap=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = Font(bold=bold, size=sz, color=color, name='Calibri', italic=italic)
    if bg: cell.fill = hfill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if num_fmt: cell.number_format = num_fmt
    cell.border = thin()
    return cell

# ─────────────────────────────────────────────────────────────────────────────
# CREATE SHEETS
# ─────────────────────────────────────────────────────────────────────────────
sheet_names = ["Dashboard","Variable_Matrix","Journal","General_Ledger",
               "Trial_Balance","Inventory_Module","PPE_Module","Lease_Module",
               "Equity_Module","Profit_Loss","SOFP","Cash_Flow",
               "Notes_Disclosures","AI_Log","Input_Form"]

ws_dash = wb.active
ws_dash.title = "Dashboard"
WS = {"Dashboard": ws_dash}
for s in sheet_names[1:]:
    WS[s] = wb.create_sheet(s)

# ─────────────────────────────────────────────────────────────────────────────
# 1. VARIABLE MATRIX
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Variable_Matrix"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "GROUP 8 — VARIABLE MATRIX", DARK_NAVY, 14)
hdr_row(ws, 2, ["Variable","Symbol","Value (PKR)","Notes"])
wcols(ws, {'A':28,'B':16,'C':18,'D':44})

vars_data = [
    ("Shares Issued",          "SHARES",      550000,   "Number of ordinary shares issued"),
    ("Share Price (PKR)",      "SHARE_PRICE",  28,       "PKR per share at issuance"),
    ("Par Value per Share",    "PAR_VALUE",    10,       "Nominal/par value per share"),
    ("PPE — Machinery Cost",   "PPE_COST",  20000000,   "IAS 16 Cost Model"),
    ("Markup %",               "MARKUP",      0.55,     "55% markup on cost"),
    ("Annual Lease Payment",   "LEASE_PMT",  450000,    "IFRS 16 annual cash payment"),
    ("Discount Rate",          "DISC_RATE",    0.13,    "13% incremental borrowing rate"),
    ("Lease Term (yrs)",       "LEASE_TERM",     5,     "5-year lease"),
    ("Research Expense",       "RESEARCH",    75000,    "IAS 38 — expensed immediately"),
    ("Development Cost",       "DEV_COST",   220000,    "IAS 38 — capitalised"),
    ("Recoverable Amount",     "RECOV_AMT",18000000,    "IAS 36 — higher of FVLCD & VIU"),
    ("Useful Life — Machinery","USEFUL_LIFE",    60,    "Years — straight-line"),
    ("Land Cost",              "LAND_COST",  5000000,   "IAS 16 land purchase"),
    ("Revaluation Gain",       "REVAL_GAIN", 1000000,   "OCI — revaluation surplus on land"),
    ("Lease PV",               "LEASE_PV",
     "=PV(C9,C10,-C8)",                                "=PV(13%,5,-450000) — IFRS 16 initial"),
]

for i, (var, sym, val, note) in enumerate(vars_data):
    r = i + 3
    c1 = ws.cell(r, 1, var); c1.font = Font(name='Calibri', sz=10)
    c2 = ws.cell(r, 2, sym); c2.font = Font(name='Calibri', size=10, bold=True, color=MID_BLUE)
    c3 = ws.cell(r, 3, val)
    is_formula = isinstance(val, str) and val.startswith('=')
    c3.font = Font(name='Calibri', size=10,
                   color=LINK_GREEN if is_formula else INPUT_BLUE)
    if isinstance(val, float) and val < 1:
        c3.number_format = '0.00%'
    else:
        c3.number_format = '#,##0'
    c4 = ws.cell(r, 4, note); c4.font = Font(name='Calibri', size=9, italic=True, color=DARK_GRAY)
    for col in range(1, 5):
        ws.cell(r, col).border = thin()
        if i % 2 == 0:
            ws.cell(r, col).fill = hfill(LIGHT_GRAY)

# Named convenience: VM row for each symbol (row = i+3)
VM = {sym: i+3 for i, (_, sym, _, _) in enumerate(vars_data)}
# VM["SHARES"]=3, VM["SHARE_PRICE"]=4, etc.

# ─────────────────────────────────────────────────────────────────────────────
# 2. JOURNAL — all 35 transactions expanded (one row per debit/credit line)
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Journal"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 9, "GENERAL JOURNAL — Group 8 | All 35 IFRS Transactions (Expanded Debit/Credit)", DARK_NAVY, 12)
hdr_row(ws, 2, ["Date","Txn ID","Description","Account","Dr/Cr","Amount (PKR)","IFRS Ref","Category","D/C Flag"], MID_BLUE)
wcols(ws, {'A':12,'B':7,'C':42,'D':34,'E':7,'F':16,'G':12,'H':14,'I':8})

# Each entry: (date, id, desc, account, dr_cr, amount_or_formula, ifrs, category)
# D/C Flag: 1=Debit, -1=Credit (used in SUMIF logic)
txns = [
    # T01 Share Issuance
    ("2024-01-01","T01","Share Issuance — 550k shares @ PKR 28","Cash","Dr",
     "=Variable_Matrix!C3*Variable_Matrix!C4","IAS 1","Equity",1),
    ("2024-01-01","T01","Share Issuance — Share Capital (par 10)",  "Share Capital","Cr",
     "=Variable_Matrix!C3*Variable_Matrix!C5","IAS 1","Equity",-1),
    ("2024-01-01","T01","Share Issuance — Share Premium",           "Share Premium","Cr",
     "=Variable_Matrix!C3*(Variable_Matrix!C4-Variable_Matrix!C5)","IAS 1","Equity",-1),
    # T02
    ("2024-01-01","T02","Legal Fees on Share Issuance","Legal Expense","Dr",50000,"IAS 1","Operating",1),
    ("2024-01-01","T02","Legal Fees on Share Issuance","Cash","Cr",50000,"IAS 1","Operating",-1),
    # T03
    ("2024-01-03","T03","Machinery Purchase","Machinery","Dr","=Variable_Matrix!C6","IAS 16","Investing",1),
    ("2024-01-03","T03","Machinery Purchase","Cash","Cr","=Variable_Matrix!C6","IAS 16","Investing",-1),
    # T04
    ("2024-01-04","T04","Land Purchase","Land","Dr","=Variable_Matrix!C15","IAS 16","Investing",1),
    ("2024-01-04","T04","Land Purchase","Cash","Cr","=Variable_Matrix!C15","IAS 16","Investing",-1),
    # T05
    ("2024-01-05","T05","Land Revaluation — OCI","Land","Dr","=Variable_Matrix!C16","IAS 16","OCI",1),
    ("2024-01-05","T05","Land Revaluation — OCI","Revaluation Surplus (OCI)","Cr","=Variable_Matrix!C16","IAS 16","OCI",-1),
    # T06
    ("2024-01-06","T06","Inventory Purchase Batch 1 (credit)","Inventory","Dr",100000,"IAS 2","Operating",1),
    ("2024-01-06","T06","Inventory Purchase Batch 1 (credit)","Accounts Payable","Cr",100000,"IAS 2","Operating",-1),
    # T07
    ("2024-01-07","T07","Sale 1 — on credit","Accounts Receivable","Dr",93000,"IAS 1","Operating",1),
    ("2024-01-07","T07","Sale 1 — on credit","Sales Revenue","Cr",93000,"IAS 1","Operating",-1),
    # T08
    ("2024-01-07","T08","COGS — Batch 1 (FIFO)","Cost of Sales","Dr",60000,"IAS 2","Operating",1),
    ("2024-01-07","T08","COGS — Batch 1 (FIFO)","Inventory","Cr",60000,"IAS 2","Operating",-1),
    # T09
    ("2024-01-08","T09","Collection from Customer","Cash","Dr",74400,"IAS 1","Operating",1),
    ("2024-01-08","T09","Collection from Customer","Accounts Receivable","Cr",74400,"IAS 1","Operating",-1),
    # T10
    ("2024-01-09","T10","Inventory Purchase Batch 2 (cash)","Inventory","Dr",60000,"IAS 2","Operating",1),
    ("2024-01-09","T10","Inventory Purchase Batch 2 (cash)","Cash","Cr",60000,"IAS 2","Operating",-1),
    # T11
    ("2024-01-10","T11","Sale 2 — cash","Cash","Dr",74400,"IAS 1","Operating",1),
    ("2024-01-10","T11","Sale 2 — cash","Sales Revenue","Cr",74400,"IAS 1","Operating",-1),
    # T12
    ("2024-01-11","T12","Sales Return","Sales Returns & Allowances","Dr",9300,"IAS 1","Operating",1),
    ("2024-01-11","T12","Sales Return","Cash","Cr",9300,"IAS 1","Operating",-1),
    # T13
    ("2024-01-12","T13","NRV Write-Down (IAS 2)","Inventory Write-Down Loss","Dr",1000,"IAS 2","Operating",1),
    ("2024-01-12","T13","NRV Write-Down (IAS 2)","Inventory","Cr",1000,"IAS 2","Operating",-1),
    # T14
    ("2024-01-13","T14","Payment to Suppliers","Accounts Payable","Dr",60000,"IAS 1","Operating",1),
    ("2024-01-13","T14","Payment to Suppliers","Cash","Cr",60000,"IAS 1","Operating",-1),
    # T15
    ("2024-01-14","T15","Bad Debt Expense","Bad Debt Expense","Dr",5000,"IAS 1","Operating",1),
    ("2024-01-14","T15","Bad Debt Expense","Accounts Receivable","Cr",5000,"IAS 1","Operating",-1),
    # T16
    ("2024-01-15","T16","Research Expense — expensed (IAS 38)","Research Expense","Dr","=Variable_Matrix!C11","IAS 38","Operating",1),
    ("2024-01-15","T16","Research Expense — expensed (IAS 38)","Cash","Cr","=Variable_Matrix!C11","IAS 38","Operating",-1),
    # T17
    ("2024-01-16","T17","Development Cost Capitalised (IAS 38)","Intangible Asset — Dev Cost","Dr","=Variable_Matrix!C12","IAS 38","Investing",1),
    ("2024-01-16","T17","Development Cost Capitalised (IAS 38)","Cash","Cr","=Variable_Matrix!C12","IAS 38","Investing",-1),
    # T18
    ("2024-01-17","T18","IFRS 16 — ROU Asset Recognition","Right-of-Use Asset","Dr","=Variable_Matrix!C17","IFRS 16","Financing",1),
    ("2024-01-17","T18","IFRS 16 — Lease Liability Recognition","Lease Liability","Cr","=Variable_Matrix!C17","IFRS 16","Financing",-1),
    # T19 — PV memo
    ("2024-01-17","T19","IFRS 16 PV Check: =PV(13%,5,-450000)","[Memo — No JE]","Dr",0,"IFRS 16","Memo",0),
    # T20
    ("2024-01-31","T20","Lease Interest — Year 1","Interest Expense","Dr",207619,"IFRS 16","Operating",1),
    ("2024-01-31","T20","Lease Interest — Year 1","Lease Liability","Cr",207619,"IFRS 16","Operating",-1),
    # T21
    ("2024-02-01","T21","Provision for Legal Claim (IAS 37)","Legal Expense — Provision","Dr",300000,"IAS 37","Operating",1),
    ("2024-02-01","T21","Provision for Legal Claim (IAS 37)","Provision Liability","Cr",300000,"IAS 37","Operating",-1),
    # T22 — disclosure only
    ("2024-02-01","T22","Contingent Liability — Disclosure Only (IAS 37)","[No JE — Disclosure]","Dr",0,"IAS 37","Memo",0),
    # T23 — assessment only
    ("2024-02-01","T23","Impairment Indicator Assessment (IAS 36) — No JE","[Assessment Only]","Dr",0,"IAS 36","Memo",0),
    # T24
    ("2024-02-01","T24","Impairment Loss on Machinery","Impairment Loss","Dr",2000000,"IAS 36","Operating",1),
    ("2024-02-01","T24","Impairment Loss on Machinery","Accum. Impairment — Machinery","Cr",2000000,"IAS 36","Operating",-1),
    # T25
    ("2024-02-01","T25","Amortisation — Dev Cost (10yr SL)","Amortisation Expense","Dr",1833,"IAS 38","Operating",1),
    ("2024-02-01","T25","Amortisation — Dev Cost (10yr SL)","Accum. Amortisation — Intangible","Cr",1833,"IAS 38","Operating",-1),
    # T26
    ("2024-02-15","T26","Treasury Share Buyback (IAS 32)","Treasury Shares","Dr",300000,"IAS 32","Equity",1),
    ("2024-02-15","T26","Treasury Share Buyback (IAS 32)","Cash","Cr",300000,"IAS 32","Equity",-1),
    # T27
    ("2024-03-01","T27","Dividend Declaration","Retained Earnings","Dr",810000,"IAS 1","Equity",1),
    ("2024-03-01","T27","Dividend Declaration","Dividend Payable","Cr",810000,"IAS 1","Equity",-1),
    # T28
    ("2024-03-15","T28","Dividend Payment","Dividend Payable","Dr",810000,"IAS 1","Equity",1),
    ("2024-03-15","T28","Dividend Payment","Cash","Cr",810000,"IAS 1","Equity",-1),
    # T29
    ("2024-03-20","T29","Bonus Issue — Capitalisation","Retained Earnings","Dr",275000,"IAS 1","Equity",1),
    ("2024-03-20","T29","Bonus Issue — Capitalisation","Share Capital","Cr",275000,"IAS 1","Equity",-1),
    # T30
    ("2024-04-01","T30","Private Placement","Cash","Dr",1650000,"IAS 1","Equity",1),
    ("2024-04-01","T30","Private Placement — Share Capital","Share Capital","Cr",500000,"IAS 1","Equity",-1),
    ("2024-04-01","T30","Private Placement — Share Premium","Share Premium","Cr",1150000,"IAS 1","Equity",-1),
    # T31
    ("2024-12-31","T31","Depreciation — Machinery (SL 60yr)","Depreciation Expense","Dr",333333,"IAS 16","Operating",1),
    ("2024-12-31","T31","Depreciation — Machinery (SL 60yr)","Accum. Depreciation — Machinery","Cr",333333,"IAS 16","Operating",-1),
    # T32
    ("2024-01-01","T32","Prepaid Insurance","Prepaid Insurance","Dr",120000,"IAS 1","Operating",1),
    ("2024-01-01","T32","Prepaid Insurance","Cash","Cr",120000,"IAS 1","Operating",-1),
    # T33
    ("2024-01-31","T33","Insurance Expense (1 month recognised)","Insurance Expense","Dr",10000,"IAS 1","Operating",1),
    ("2024-01-31","T33","Insurance Expense (1 month recognised)","Prepaid Insurance","Cr",10000,"IAS 1","Operating",-1),
    # T34
    ("2024-12-31","T34","Utilities Accrual","Utilities Expense","Dr",20000,"IAS 1","Operating",1),
    ("2024-12-31","T34","Utilities Accrual","Utilities Payable","Cr",20000,"IAS 1","Operating",-1),
    # T35
    ("2024-12-31","T35","Salaries Expense","Salaries Expense","Dr",150000,"IAS 1","Operating",1),
    ("2024-12-31","T35","Salaries Expense","Cash","Cr",150000,"IAS 1","Operating",-1),
]

DATA_START = 3
cat_colors = {"Dr": ACCENT_RED, "Cr": ACCENT_GREEN, "Dr/Cr": DARK_GRAY}

for i, txn in enumerate(txns):
    r = DATA_START + i
    ws.row_dimensions[r].height = 17
    date_, tid, desc, acct, dc, amt, ifrs, cat, flag = txn
    vals = [date_, tid, desc, acct, dc, amt, ifrs, cat, flag]
    for col, val in enumerate(vals, 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='left', vertical='center')
        if i % 2 == 0:
            c.fill = hfill(LIGHT_GRAY)
        if col == 5:  # Dr/Cr
            c.font = Font(name='Calibri', size=9, bold=True,
                          color=ACCENT_RED if dc == 'Dr' else (ACCENT_GREEN if dc == 'Cr' else DARK_GRAY))
            c.alignment = Alignment(horizontal='center', vertical='center')
        if col == 6:
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal='right', vertical='center')
            if isinstance(val, str) and val.startswith('='):
                c.font = Font(name='Calibri', size=9, color=LINK_GREEN)
        if col == 9:
            c.alignment = Alignment(horizontal='center', vertical='center')

JNRL_END = DATA_START + len(txns) - 1

# Totals
tr = JNRL_END + 1
ws.cell(tr, 3, "TOTAL DEBITS / CREDITS").font = Font(bold=True, name='Calibri')
ws.cell(tr, 6, f"=SUMPRODUCT((I{DATA_START}:I{JNRL_END}=1)*F{DATA_START}:F{JNRL_END})")
ws.cell(tr, 6).number_format = '#,##0'
ws.cell(tr, 6).fill = hfill(ACCENT_GOLD)
ws.cell(tr, 6).font = Font(bold=True, name='Calibri')

# Balance check
br = tr + 1
ws.cell(br, 3, "Balance Check (Dr = Cr):").font = Font(bold=True, name='Calibri')
formula = (f'=IF(ABS(SUMPRODUCT((I{DATA_START}:I{JNRL_END}=1)*F{DATA_START}:F{JNRL_END})'
           f'-SUMPRODUCT((I{DATA_START}:I{JNRL_END}=-1)*F{DATA_START}:F{JNRL_END}))<1,'
           f'"✔ BALANCED","✘ CHECK ENTRIES")')
ws.cell(br, 6, formula)
ws.cell(br, 6).font = Font(bold=True, color=ACCENT_GREEN, name='Calibri')

# ─────────────────────────────────────────────────────────────────────────────
# 3. GENERAL LEDGER
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["General_Ledger"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 7, "GENERAL LEDGER — Auto-Posting from Journal via SUMIF", DARK_NAVY, 13)
hdr_row(ws, 2, ["Account","Type","Class","Total Debits","Total Credits","Net Balance","Normal Side"])
wcols(ws, {'A':38,'B':14,'C':14,'D':18,'E':18,'F':18,'G':12})

# Account master list — (name, type, normal_balance)
accounts = [
    ("Cash",                            "Asset",       "Dr"),
    ("Accounts Receivable",             "Asset",       "Dr"),
    ("Inventory",                       "Asset",       "Dr"),
    ("Prepaid Insurance",               "Asset",       "Dr"),
    ("Machinery",                       "Asset",       "Dr"),
    ("Land",                            "Asset",       "Dr"),
    ("Right-of-Use Asset",              "Asset",       "Dr"),
    ("Intangible Asset — Dev Cost",     "Asset",       "Dr"),
    ("Accum. Depreciation — Machinery", "Contra",      "Cr"),
    ("Accum. Impairment — Machinery",   "Contra",      "Cr"),
    ("Accum. Amortisation — Intangible","Contra",      "Cr"),
    ("Accounts Payable",                "Liability",   "Cr"),
    ("Utilities Payable",               "Liability",   "Cr"),
    ("Dividend Payable",                "Liability",   "Cr"),
    ("Lease Liability",                 "Liability",   "Cr"),
    ("Provision Liability",             "Liability",   "Cr"),
    ("Share Capital",                   "Equity",      "Cr"),
    ("Share Premium",                   "Equity",      "Cr"),
    ("Revaluation Surplus (OCI)",       "Equity",      "Cr"),
    ("Retained Earnings",               "Equity",      "Cr"),
    ("Treasury Shares",                 "Contra-Eq",   "Dr"),
    ("Sales Revenue",                   "Revenue",     "Cr"),
    ("Sales Returns & Allowances",      "Expense",     "Dr"),
    ("Cost of Sales",                   "Expense",     "Dr"),
    ("Inventory Write-Down Loss",       "Expense",     "Dr"),
    ("Legal Expense",                   "Expense",     "Dr"),
    ("Legal Expense — Provision",       "Expense",     "Dr"),
    ("Bad Debt Expense",                "Expense",     "Dr"),
    ("Research Expense",                "Expense",     "Dr"),
    ("Depreciation Expense",            "Expense",     "Dr"),
    ("Amortisation Expense",            "Expense",     "Dr"),
    ("Impairment Loss",                 "Expense",     "Dr"),
    ("Interest Expense",                "Expense",     "Dr"),
    ("Insurance Expense",               "Expense",     "Dr"),
    ("Utilities Expense",               "Expense",     "Dr"),
    ("Salaries Expense",                "Expense",     "Dr"),
]

GL_DATA_START = 3
# Build a dict: account_name -> GL row
GL_ROW = {name: GL_DATA_START + i for i, (name, _, _) in enumerate(accounts)}

J = "Journal"
for i, (acct, typ, norm) in enumerate(accounts):
    r = GL_DATA_START + i
    # Debits: rows in Journal where Account=this AND flag=1
    deb = f'=SUMPRODUCT((Journal!D{DATA_START}:D{JNRL_END}=A{r})*(Journal!I{DATA_START}:I{JNRL_END}=1)*Journal!F{DATA_START}:F{JNRL_END})'
    crd = f'=SUMPRODUCT((Journal!D{DATA_START}:D{JNRL_END}=A{r})*(Journal!I{DATA_START}:I{JNRL_END}=-1)*Journal!F{DATA_START}:F{JNRL_END})'
    if norm == "Dr":
        net = f'=D{r}-E{r}'
    else:
        net = f'=E{r}-D{r}'
    row_data = [acct, typ, norm+"-side", deb, crd, net, norm]
    for col, val in enumerate(row_data, 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=9, bold=(col==1 and typ in ("Asset","Liability","Equity")))
        c.alignment = Alignment(horizontal='right' if col > 3 else 'left', vertical='center')
        if col in (4, 5, 6): c.number_format = '#,##0'
        if i % 2 == 0: c.fill = hfill(LIGHT_GRAY)

# ─────────────────────────────────────────────────────────────────────────────
# 4. TRIAL BALANCE
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Trial_Balance"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 5, "TRIAL BALANCE — As at 31 December 2024", DARK_NAVY, 13)
hdr_row(ws, 2, ["Account Name","Type","Debit (PKR)","Credit (PKR)","Source"])
wcols(ws, {'A':38,'B':14,'C':18,'D':18,'E':22})

TB_START = 3
for i, (acct, typ, norm) in enumerate(accounts):
    r = TB_START + i
    gl_r = GL_ROW[acct]
    ws.cell(r, 1, acct).font = Font(name='Calibri', size=9)
    ws.cell(r, 2, typ).font = Font(name='Calibri', size=9)
    net_ref = f"=General_Ledger!F{gl_r}"
    if norm == "Dr":
        ws.cell(r, 3, f"=IF(General_Ledger!F{gl_r}>0,General_Ledger!F{gl_r},0)")
        ws.cell(r, 4, "")
    else:
        ws.cell(r, 3, "")
        ws.cell(r, 4, f"=IF(General_Ledger!F{gl_r}>0,General_Ledger!F{gl_r},0)")
    ws.cell(r, 5, acct)
    for col in range(1, 6):
        c = ws.cell(r, col)
        c.border = thin()
        c.font = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
        if col in (3, 4): c.number_format = '#,##0'
        if i % 2 == 0: c.fill = hfill(LIGHT_GRAY)

# Total row
tb_total = TB_START + len(accounts)
hdr_row(ws, tb_total, ["TOTALS","","","",""], ACCENT_GOLD)
ws.cell(tb_total, 3, f"=SUM(C{TB_START}:C{tb_total-1})").number_format = '#,##0'
ws.cell(tb_total, 4, f"=SUM(D{TB_START}:D{tb_total-1})").number_format = '#,##0'
for col in [3, 4]:
    ws.cell(tb_total, col).font = Font(bold=True, name='Calibri')
    ws.cell(tb_total, col).fill = hfill(ACCENT_GOLD)

chk = tb_total + 2
ws.merge_cells(f'A{chk}:B{chk}')
ws.cell(chk, 1, "Balance Check:").font = Font(bold=True, name='Calibri')
ws.cell(chk, 3, f'=IF(ABS(C{tb_total}-D{tb_total})<1,"✔ BALANCED","✘ ERROR")').font = Font(bold=True, color=ACCENT_GREEN, name='Calibri')

# ─────────────────────────────────────────────────────────────────────────────
# 5. INVENTORY MODULE
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Inventory_Module"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 8, "INVENTORY MODULE — IAS 2 | FIFO Costing & NRV Test", DARK_NAVY, 13)
hdr_row(ws, 2, ["Batch","Date","Units In","Cost/Unit (PKR)","Total Cost","Units Sold","COGS","Closing Units"])
wcols(ws, {'A':10,'B':14,'C':12,'D':16,'E':16,'F':14,'G':16,'H':16})

inv_rows = [
    ("B1","2024-01-06",100,1000,"=C3*D3",60,"=F3*D3","=C3-F3"),
    ("B2","2024-01-09", 60,1000,"=C4*D4",40,"=F4*D4","=C4-F4"),
]
for i, row in enumerate(inv_rows):
    r = i + 3
    for col, val in enumerate(row, 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=10, color=LINK_GREEN if isinstance(val, str) and val.startswith('=') else INPUT_BLUE)
        c.alignment = Alignment(horizontal='right' if col > 2 else 'center', vertical='center')
        if col in (4,5,7): c.number_format = '#,##0'

# NRV section
title(ws, 7, 1, 8, "NRV ANALYSIS — IAS 2.9", ACCENT_RED, 11, h=22)
nrv_items = [
    ("Closing Inventory — Cost:",   "=H3*D3+H4*D4", None),
    ("Net Realisable Value (NRV):", 99000,           "IAS 2 estimate"),
    ("Write-Down Required:",        "=MAX(0,C8-C9)", "= MAX(0, Cost - NRV)"),
]
for i, (label, val, note) in enumerate(nrv_items):
    r = i + 8
    ws.cell(r, 1, label).font = Font(bold=True, name='Calibri', size=10)
    c = ws.cell(r, 3, val)
    c.number_format = '#,##0'
    c.font = Font(name='Calibri', size=10,
                  color=LINK_GREEN if isinstance(val, str) and val.startswith('=') else INPUT_BLUE)
    if note:
        ws.cell(r, 4, note).font = Font(italic=True, color=DARK_GRAY, name='Calibri', size=9)

title(ws, 13, 1, 8, "SUMMARY", MID_BLUE, 10, h=20)
ws.cell(14, 1, "Total COGS:").font = Font(bold=True, name='Calibri')
ws.cell(14, 3, "=G3+G4").number_format = '#,##0'
ws.cell(15, 1, "Closing Inventory (after write-down):").font = Font(bold=True, name='Calibri')
ws.cell(15, 3, "=C8-C10").number_format = '#,##0'

# ─────────────────────────────────────────────────────────────────────────────
# 6. PPE MODULE
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["PPE_Module"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 7, "PPE MODULE — IAS 16 | IAS 36 Impairment", DARK_NAVY, 13)
hdr_row(ws, 2, ["Asset","Cost","Accum. Depr.","Impairment","Carrying Amt","Recoverable","Imp. Loss"])
wcols(ws, {'A':22,'B':18,'C':18,'D':18,'E':18,'F':18,'G':18})

ppe_rows = [
    ("Machinery",
     "=Variable_Matrix!C6",
     "=Variable_Matrix!C6/Variable_Matrix!C14",
     2000000,
     "=B3-C3-D3",
     "=Variable_Matrix!C13",
     "=MAX(0,B3-D3-F3)"),
    ("Land",
     "=Variable_Matrix!C15+Variable_Matrix!C16",
     0, 0,
     "=B4",
     "N/A",
     0),
]
for i, row in enumerate(ppe_rows):
    r = i + 3
    for col, val in enumerate(row, 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=10)
        c.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center')
        if col > 1 and col != 6:
            c.number_format = '#,##0'

ws.cell(6, 1, "Revaluation Surplus (Land):").font = Font(bold=True, name='Calibri')
ws.cell(6, 3, "=Variable_Matrix!C16").number_format = '#,##0'
ws.cell(6, 3).font = Font(bold=True, color=LINK_GREEN, name='Calibri')

# Depreciation schedule
title(ws, 8, 1, 7, "STRAIGHT-LINE DEPRECIATION SCHEDULE — Machinery", MID_BLUE, 11, h=22)
hdr_row(ws, 9, ["Year","Opening CA","Annual Depr.","Impairment","Closing CA","Accum Depr","Accum Imp"], ACCENT_BLUE)
for yr in range(1, 4):
    r = yr + 9
    op   = "=Variable_Matrix!C6" if yr == 1 else f"=E{r-1}"
    depr = "=Variable_Matrix!C6/Variable_Matrix!C14"
    imp  = 2000000 if yr == 1 else 0
    cls  = f"=B{r}-C{r}-D{r}"
    ad   = f"=SUM(C$10:C{r})"
    ai   = f"=SUM(D$10:D{r})"
    for col, val in enumerate([yr, op, depr, imp, cls, ad, ai], 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=10, color=LINK_GREEN if isinstance(val,str) and val.startswith('=') else BLACK)
        if col > 1: c.number_format = '#,##0'
        if yr % 2 == 0: c.fill = hfill(LIGHT_GRAY)

# ─────────────────────────────────────────────────────────────────────────────
# 7. LEASE MODULE
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Lease_Module"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 7, "LEASE MODULE — IFRS 16 | Amortisation Schedule", DARK_NAVY, 13)
wcols(ws, {'A':30,'B':4,'C':20,'D':18,'E':18,'F':20,'G':18})

params = [
    ("Annual Lease Payment (PKR):",   "=Variable_Matrix!C8"),
    ("Discount Rate:",                 "=Variable_Matrix!C9"),
    ("Lease Term (Years):",            "=Variable_Matrix!C10"),
    ("PV of Lease Liability (ROU):",   "=PV(C3,C4,-C2)"),
]
for i, (lbl, formula) in enumerate(params):
    r = i + 2
    ws.cell(r, 1, lbl).font = Font(bold=True, name='Calibri', size=10)
    c = ws.cell(r, 3, formula)
    c.font = Font(bold=(i==3), name='Calibri', size=10,
                  color=LINK_GREEN if formula.startswith('=') else INPUT_BLUE)
    c.number_format = '0.00%' if i == 1 else '#,##0'

ws.cell(5, 4, "← IFRS 16 initial measurement basis").font = Font(italic=True, color=DARK_GRAY, name='Calibri', size=9)

hdr_row(ws, 7, ["Year","Opening Liability","Interest (13%)","Lease Payment","Principal","Closing Liability","ROU Depr"], MID_BLUE)
for yr in range(1, 6):
    r = yr + 7
    op      = "=C5" if yr == 1 else f"=F{r-1}"
    intr    = f"=B{r}*C$3"
    pmt     = "=C$2"
    princ   = f"=D{r}-C{r}"
    closing = f"=B{r}-E{r}"
    rou_d   = "=C$5/C$4"
    for col, val in enumerate([yr, op, intr, pmt, princ, closing, rou_d], 1):
        c = ws.cell(r, col, val)
        c.number_format = '#,##0' if col > 1 else '0'
        c.border = thin()
        c.font = Font(name='Calibri', size=10, color=LINK_GREEN if isinstance(val,str) and val.startswith('=') else BLACK)
        c.alignment = Alignment(horizontal='right' if col > 1 else 'center', vertical='center')
        if yr % 2 == 0: c.fill = hfill(LIGHT_GRAY)

# Totals
tr2 = 13
ws.cell(tr2, 1, "Totals").font = Font(bold=True, name='Calibri')
for col, f in [(3,"=SUM(C8:C12)"),(4,"=SUM(D8:D12)"),(5,"=SUM(E8:E12)"),(7,"=SUM(G8:G12)")]:
    c = ws.cell(tr2, col, f)
    c.number_format = '#,##0'
    c.font = Font(bold=True, name='Calibri')
    c.fill = hfill(ACCENT_GOLD)

# ─────────────────────────────────────────────────────────────────────────────
# 8. EQUITY MODULE
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Equity_Module"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 7, "STATEMENT OF CHANGES IN EQUITY — Year Ended 31 December 2024", DARK_NAVY, 13)
hdr_row(ws, 2, ["Item","Share Capital","Share Premium","Ret. Earnings","Reval. Surplus","Treasury Shares","Total Equity"])
wcols(ws, {'A':34,'B':16,'C':16,'D':16,'E':16,'F':16,'G':16})

# GL rows for equity accounts
sc_r  = GL_ROW["Share Capital"]
sp_r  = GL_ROW["Share Premium"]
re_r  = GL_ROW["Retained Earnings"]
oc_r  = GL_ROW["Revaluation Surplus (OCI)"]
ts_r  = GL_ROW["Treasury Shares"]

eq_rows = [
    ("Opening Balance (1 Jan 2024)",   0,0,0,0,0),
    ("T01 — Share Issuance",
     f"=General_Ledger!E{sc_r}-General_Ledger!D{sc_r}",
     f"=General_Ledger!E{sp_r}-General_Ledger!D{sp_r}",
     0,0,0),
    ("T05 — Revaluation Surplus",      0,0,0,"=Variable_Matrix!C16",0),
    ("T27/T28 — Dividends",            0,0,-810000,0,0),
    ("T29 — Bonus Issue",              275000,0,-275000,0,0),
    ("T30 — Private Placement",        500000,1150000,0,0,0),
    ("T26 — Treasury Shares",          0,0,0,0,-300000),
    ("Net Profit for Year",            0,0,"=Profit_Loss!C37",0,0),
    ("Closing Balance (31 Dec 2024)",
     f"=General_Ledger!F{sc_r}",
     f"=General_Ledger!F{sp_r}",
     f"=General_Ledger!F{re_r}",
     f"=General_Ledger!F{oc_r}",
     f"=-General_Ledger!F{ts_r}"),
]

for i, row in enumerate(eq_rows):
    r = i + 3
    ws.row_dimensions[r].height = 18
    is_close = i == len(eq_rows) - 1
    is_open  = i == 0
    for col, val in enumerate(row, 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=10, bold=(is_close or is_open),
                      color=LINK_GREEN if isinstance(val, str) and val.startswith('=') else (BLACK if isinstance(val, (int, float)) else BLACK))
        c.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center')
        c.number_format = '#,##0'
        if is_close: c.fill = hfill(LIGHT_BLUE)
    # Total column
    c = ws.cell(r, 7, f"=SUM(B{r}:F{r})")
    c.border = thin()
    c.font = Font(name='Calibri', size=10, bold=(is_close or is_open))
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal='right', vertical='center')
    if is_close: c.fill = hfill(LIGHT_BLUE)

# ─────────────────────────────────────────────────────────────────────────────
# 9. PROFIT & LOSS
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Profit_Loss"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "STATEMENT OF PROFIT OR LOSS — Year Ended 31 December 2024", DARK_NAVY, 13)
title(ws, 2, 1, 4, "Prepared in Accordance with IAS 1 (Revised)", MID_BLUE, 10, h=18)
hdr_row(ws, 3, ["Description","IFRS Ref","Amount (PKR)","Notes"])
wcols(ws, {'A':38,'B':12,'C':20,'D':28})

GL = "General_Ledger"
gl_r = GL_ROW  # shorthand

pl_items = [
    # (desc, ref, formula_or_val, note, is_section, is_total)
    ("REVENUE","","","",True,False),
    ("Sales Revenue","IAS 1",f"=General_Ledger!F{gl_r['Sales Revenue']}","T07+T11",False,False),
    ("Less: Sales Returns","IAS 1",f"=-General_Ledger!F{gl_r['Sales Returns & Allowances']}","T12",False,False),
    ("Net Revenue","","=C5+C6","",False,True),
    ("","","","",False,False),
    ("COST OF SALES","","","",True,False),
    ("Cost of Goods Sold","IAS 2",f"=-General_Ledger!F{gl_r['Cost of Sales']}","T08",False,False),
    ("Inventory Write-Down","IAS 2",f"=-General_Ledger!F{gl_r['Inventory Write-Down Loss']}","T13",False,False),
    ("Total Cost of Sales","","=-(C9+C10)","",False,True),
    ("","","","",False,False),
    ("GROSS PROFIT","","=C7+C11","",False,True),
    ("","","","",False,False),
    ("OPERATING EXPENSES","","","",True,False),
    ("Depreciation — Machinery","IAS 16",f"=-General_Ledger!F{gl_r['Depreciation Expense']}","T31",False,False),
    ("Amortisation — Dev Cost","IAS 38",f"=-General_Ledger!F{gl_r['Amortisation Expense']}","T25",False,False),
    ("Impairment Loss","IAS 36",f"=-General_Ledger!F{gl_r['Impairment Loss']}","T24",False,False),
    ("Research Expense","IAS 38",f"=-General_Ledger!F{gl_r['Research Expense']}","T16",False,False),
    ("Legal Expense","IAS 37",f"=-(General_Ledger!F{gl_r['Legal Expense']}+General_Ledger!F{gl_r['Legal Expense — Provision']})","T02+T21",False,False),
    ("Bad Debt Expense","IAS 1",f"=-General_Ledger!F{gl_r['Bad Debt Expense']}","T15",False,False),
    ("Insurance Expense","IAS 1",f"=-General_Ledger!F{gl_r['Insurance Expense']}","T33",False,False),
    ("Utilities Expense","IAS 1",f"=-General_Ledger!F{gl_r['Utilities Expense']}","T34",False,False),
    ("Salaries Expense","IAS 1",f"=-General_Ledger!F{gl_r['Salaries Expense']}","T35",False,False),
    ("Total Operating Expenses","","=SUM(C15:C25)","",False,True),
    ("","","","",False,False),
    ("OPERATING PROFIT (EBIT)","","=C12+C26","",False,True),
    ("","","","",False,False),
    ("FINANCE COSTS","","","",True,False),
    ("Interest Expense — Lease","IFRS 16",f"=-General_Ledger!F{gl_r['Interest Expense']}","T20",False,False),
    ("","","","",False,False),
    ("PROFIT BEFORE TAX","","=C28+C31","",False,True),
    ("Income Tax Expense","IAS 12","=0","",False,False),
    ("NET PROFIT FOR THE YEAR","","=C32+C33","",False,True),
    ("","","","",False,False),
    ("OTHER COMPREHENSIVE INCOME","","","",True,False),
    ("Revaluation Surplus — Land","IAS 16",f"=General_Ledger!F{gl_r['Revaluation Surplus (OCI)']}","T05",False,False),
    ("TOTAL COMPREHENSIVE INCOME","","=C34+C38","",False,True),
]

for i, (desc, ref, amt, note, is_sec, is_tot) in enumerate(pl_items):
    r = i + 4
    ws.row_dimensions[r].height = 17
    c1 = ws.cell(r, 1, desc)
    c1.font = Font(name='Calibri', size=10, bold=(is_tot or is_sec),
                   color=(WHITE if is_sec else (DARK_NAVY if is_tot else BLACK)))
    if is_sec:
        c1.fill = hfill(MID_BLUE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    else:
        if is_tot:
            c1.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 2, ref).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        c3 = ws.cell(r, 3, amt if amt else "")
        c3.number_format = '#,##0'
        c3.font = Font(name='Calibri', size=10, bold=is_tot,
                       color=LINK_GREEN if isinstance(amt, str) and amt.startswith('=') else INPUT_BLUE)
        c3.alignment = Alignment(horizontal='right', vertical='center')
        if is_tot: c3.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 4, note).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        for col in range(1, 5):
            ws.cell(r, col).border = thin()

# Remember: Net Profit is at row 34+3 = row 37 (i=30 → r=34 in pl_items index)
# Let's note the actual row of NET PROFIT row:
# item index 30 (0-based) = r = 30+4 = 34
NP_ROW = None
for i, (desc, *_) in enumerate(pl_items):
    if desc == "NET PROFIT FOR THE YEAR":
        NP_ROW = i + 4
        break

# ─────────────────────────────────────────────────────────────────────────────
# 10. SOFP
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["SOFP"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "STATEMENT OF FINANCIAL POSITION — As at 31 December 2024", DARK_NAVY, 13)
title(ws, 2, 1, 4, "Prepared in Accordance with IAS 1", MID_BLUE, 10, h=16)
hdr_row(ws, 3, ["Description","Note","Amount (PKR)","IFRS Ref"])
wcols(ws, {'A':38,'B':6,'C':22,'D':16})

mac_r = GL_ROW["Machinery"]
land_r = GL_ROW["Land"]
rou_r  = GL_ROW["Right-of-Use Asset"]
int_r  = GL_ROW["Intangible Asset — Dev Cost"]
ad_r   = GL_ROW["Accum. Depreciation — Machinery"]
ai_r   = GL_ROW["Accum. Impairment — Machinery"]
aa_r   = GL_ROW["Accum. Amortisation — Intangible"]
cash_r = GL_ROW["Cash"]
ar_r   = GL_ROW["Accounts Receivable"]
inv_r  = GL_ROW["Inventory"]
prep_r = GL_ROW["Prepaid Insurance"]
ap_r   = GL_ROW["Accounts Payable"]
up_r   = GL_ROW["Utilities Payable"]
dp_r   = GL_ROW["Dividend Payable"]
ll_r   = GL_ROW["Lease Liability"]
pv_r   = GL_ROW["Provision Liability"]
sc_r2  = GL_ROW["Share Capital"]
sp_r2  = GL_ROW["Share Premium"]
re_r2  = GL_ROW["Retained Earnings"]
oc_r2  = GL_ROW["Revaluation Surplus (OCI)"]
ts_r2  = GL_ROW["Treasury Shares"]

sofp_items = [
    ("NON-CURRENT ASSETS","","","",True,False),
    ("Machinery — Cost","","=General_Ledger!F{mac_r}","IAS 16",False,False),
    ("Less: Accum. Depreciation","","=-General_Ledger!F{ad_r}","IAS 16",False,False),
    ("Less: Accum. Impairment","","=-General_Ledger!F{ai_r}","IAS 36",False,False),
    ("Machinery — Net Carrying Amount","","=C5+C6+C7","",False,True),
    ("Land (incl. Revaluation)","","=General_Ledger!F{land_r}","IAS 16",False,False),
    ("Right-of-Use Asset — Cost","","=General_Ledger!F{rou_r}","IFRS 16",False,False),
    ("Less: ROU Accumulated Depr.","","=-Lease_Module!G8","IFRS 16",False,False),
    ("ROU Asset — Net","","=C10+C11","",False,True),
    ("Intangible Asset — Dev Cost","","=General_Ledger!F{int_r}","IAS 38",False,False),
    ("Less: Accum. Amortisation","","=-General_Ledger!F{aa_r}","IAS 38",False,False),
    ("Intangible — Net","","=C14+C15","",False,True),
    ("TOTAL NON-CURRENT ASSETS","","=C8+C9+C12+C16","",False,True),
    ("","","","",False,False),
    ("CURRENT ASSETS","","","",True,False),
    ("Inventories","","=General_Ledger!F{inv_r}","IAS 2",False,False),
    ("Accounts Receivable (net)","","=General_Ledger!F{ar_r}","IAS 1",False,False),
    ("Prepaid Insurance","","=General_Ledger!F{prep_r}","IAS 1",False,False),
    ("Cash and Cash Equivalents","","=General_Ledger!F{cash_r}","IAS 7",False,False),
    ("TOTAL CURRENT ASSETS","","=SUM(C20:C23)","",False,True),
    ("","","","",False,False),
    ("TOTAL ASSETS","","=C17+C24","",False,True),
    ("","","","",False,False),
    ("EQUITY","","","",True,False),
    ("Share Capital","","=General_Ledger!F{sc_r2}","IAS 1",False,False),
    ("Share Premium","","=General_Ledger!F{sp_r2}","IAS 1",False,False),
    ("Retained Earnings","","=General_Ledger!F{re_r2}","IAS 1",False,False),
    ("Revaluation Surplus (OCI)","","=General_Ledger!F{oc_r2}","IAS 16",False,False),
    ("Less: Treasury Shares","","=-General_Ledger!F{ts_r2}","IAS 32",False,False),
    ("TOTAL EQUITY","","=SUM(C28:C32)","",False,True),
    ("","","","",False,False),
    ("NON-CURRENT LIABILITIES","","","",True,False),
    ("Lease Liability (non-current)","","=General_Ledger!F{ll_r}","IFRS 16",False,False),
    ("Provision Liability","","=General_Ledger!F{pv_r}","IAS 37",False,False),
    ("TOTAL NON-CURRENT LIABILITIES","","=C37+C38","",False,True),
    ("","","","",False,False),
    ("CURRENT LIABILITIES","","","",True,False),
    ("Accounts Payable","","=General_Ledger!F{ap_r}","IAS 1",False,False),
    ("Utilities Payable","","=General_Ledger!F{up_r}","IAS 1",False,False),
    ("Dividend Payable","","=General_Ledger!F{dp_r}","IAS 1",False,False),
    ("TOTAL CURRENT LIABILITIES","","=SUM(C43:C45)","",False,True),
    ("","","","",False,False),
    ("TOTAL LIABILITIES","","=C39+C46","",False,True),
    ("TOTAL EQUITY AND LIABILITIES","","=C33+C48","",False,True),
    ("","","","",False,False),
    ("ACCOUNTING EQUATION CHECK","","","",True,False),
    ('=IF(ABS(C26-(C33+C48))<1,"✔ Assets = Equity + Liabilities  — BALANCED","✘ EQUATION ERROR")',
     "","","",False,False),
]

# Replace placeholder refs with actual GL row numbers
def fix_formula(f):
    if not isinstance(f, str): return f
    return (f.replace("{mac_r}", str(mac_r))
             .replace("{ad_r}", str(ad_r))
             .replace("{ai_r}", str(ai_r))
             .replace("{land_r}", str(land_r))
             .replace("{rou_r}", str(rou_r))
             .replace("{int_r}", str(int_r))
             .replace("{aa_r}", str(aa_r))
             .replace("{inv_r}", str(inv_r))
             .replace("{ar_r}", str(ar_r))
             .replace("{prep_r}", str(prep_r))
             .replace("{cash_r}", str(cash_r))
             .replace("{ap_r}", str(ap_r))
             .replace("{up_r}", str(up_r))
             .replace("{dp_r}", str(dp_r))
             .replace("{ll_r}", str(ll_r))
             .replace("{pv_r}", str(pv_r))
             .replace("{sc_r2}", str(sc_r2))
             .replace("{sp_r2}", str(sp_r2))
             .replace("{re_r2}", str(re_r2))
             .replace("{oc_r2}", str(oc_r2))
             .replace("{ts_r2}", str(ts_r2)))

for i, (desc, note, amt, ref, is_sec, is_tot) in enumerate(sofp_items):
    r = i + 4
    ws.row_dimensions[r].height = 16
    amt_fixed = fix_formula(amt)
    desc_fixed = fix_formula(desc)
    is_tot = is_tot or ("TOTAL" in desc_fixed and not is_sec)
    c1 = ws.cell(r, 1, desc_fixed)
    c1.font = Font(name='Calibri', size=10, bold=(is_tot or is_sec),
                   color=WHITE if is_sec else (DARK_NAVY if is_tot else BLACK))
    if is_sec:
        c1.fill = hfill(MID_BLUE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    else:
        if is_tot: c1.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 2, note).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        c3 = ws.cell(r, 3, amt_fixed if amt_fixed else "")
        c3.number_format = '#,##0'
        c3.font = Font(name='Calibri', size=10, bold=is_tot,
                       color=LINK_GREEN if isinstance(amt_fixed, str) and amt_fixed.startswith('=') else BLACK)
        c3.alignment = Alignment(horizontal='right', vertical='center')
        if is_tot: c3.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 4, ref).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        for col in range(1, 5):
            ws.cell(r, col).border = thin()

# Equation check row gets special font
# (it's the last item, index len-1, row = len+4-1)
eq_row = len(sofp_items) + 4 - 1
ws.cell(eq_row, 1).font = Font(bold=True, size=11, color=ACCENT_GREEN, name='Calibri')
ws.cell(eq_row, 1).fill = hfill(LIGHT_GRAY)

# ─────────────────────────────────────────────────────────────────────────────
# 11. CASH FLOW
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Cash_Flow"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "STATEMENT OF CASH FLOWS — Year Ended 31 December 2024", DARK_NAVY, 13)
title(ws, 2, 1, 4, "Indirect Method — IAS 7", MID_BLUE, 10, h=16)
hdr_row(ws, 3, ["Description","Ref","Amount (PKR)","IAS 7 Note"])
wcols(ws, {'A':40,'B':8,'C':22,'D':28})

cf_items = [
    ("OPERATING ACTIVITIES","","","",True,False),
    ("Net Profit for the Year","PL",f"=Profit_Loss!C{NP_ROW}","Start: Net Profit",False,False),
    ("Add: Depreciation (non-cash)","T31",f"=ABS(General_Ledger!F{gl_r['Depreciation Expense']})","Non-cash add-back",False,False),
    ("Add: Amortisation (non-cash)","T25",f"=ABS(General_Ledger!F{gl_r['Amortisation Expense']})","Non-cash add-back",False,False),
    ("Add: Impairment (non-cash)","T24",f"=ABS(General_Ledger!F{gl_r['Impairment Loss']})","Non-cash add-back",False,False),
    ("Add: Bad Debt (non-cash)","T15",f"=ABS(General_Ledger!F{gl_r['Bad Debt Expense']})","Non-cash add-back",False,False),
    ("Working Capital Changes:","","","",False,False),
    ("(Inc)/Dec in Receivables","",f"=-General_Ledger!F{gl_r['Accounts Receivable']}","Current asset Δ",False,False),
    ("(Inc)/Dec in Inventories","",f"=-General_Ledger!F{gl_r['Inventory']}","Current asset Δ",False,False),
    ("(Inc)/Dec in Prepayments","",f"=-General_Ledger!F{gl_r['Prepaid Insurance']}","Current asset Δ",False,False),
    ("Inc/(Dec) in Payables","",f"=General_Ledger!F{gl_r['Accounts Payable']}","Current liability Δ",False,False),
    ("Inc/(Dec) in Utilities Payable","",f"=General_Ledger!F{gl_r['Utilities Payable']}","Current liability Δ",False,False),
    ("NET CASH FROM OPERATING ACTIVITIES","","=SUM(C5:C16)","",False,True),
    ("","","","",False,False),
    ("INVESTING ACTIVITIES","","","",True,False),
    ("Purchase of Machinery","T03","=-Variable_Matrix!C6","IAS 7",False,False),
    ("Purchase of Land","T04","=-Variable_Matrix!C15","IAS 7",False,False),
    ("Development Costs Capitalised","T17","=-Variable_Matrix!C12","IAS 38",False,False),
    ("NET CASH USED IN INVESTING","","=SUM(C20:C22)","",False,True),
    ("","","","",False,False),
    ("FINANCING ACTIVITIES","","","",True,False),
    ("Proceeds — Share Issuance","T01","=Variable_Matrix!C3*Variable_Matrix!C4","IAS 7",False,False),
    ("Proceeds — Private Placement","T30","=1650000","IAS 7",False,False),
    ("Treasury Share Buyback","T26","=-300000","IAS 7",False,False),
    ("Dividends Paid","T28","=-810000","IAS 7",False,False),
    ("Lease Payment — Principal","T18","=-Lease_Module!E8","IFRS 16",False,False),
    ("NET CASH FROM FINANCING","","=SUM(C25:C29)","",False,True),
    ("","","","",False,False),
    ("NET INC/(DEC) IN CASH","","=C17+C23+C30","",False,True),
    ("Opening Cash Balance","","=0","",False,False),
    ("CLOSING CASH BALANCE","","=C32+C33","",False,True),
    ("GL Cash Verification","",f'=IF(ABS(C34-General_Ledger!F{gl_r["Cash"]})<1,"✔ MATCHES GL","✘ DISCREPANCY")',"",False,False),
]

for i, (desc, ref, amt, note, is_sec, is_tot) in enumerate(cf_items):
    r = i + 4
    ws.row_dimensions[r].height = 17
    is_tot2 = is_tot or ("NET CASH" in desc or "CLOSING" in desc or "NET INC" in desc)
    c1 = ws.cell(r, 1, desc)
    c1.font = Font(name='Calibri', size=10, bold=(is_tot2 or is_sec),
                   color=WHITE if is_sec else (DARK_NAVY if is_tot2 else BLACK))
    if is_sec:
        c1.fill = hfill(MID_BLUE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    else:
        if is_tot2: c1.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 2, ref).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        c3 = ws.cell(r, 3, amt if amt else "")
        c3.number_format = '#,##0'
        c3.font = Font(name='Calibri', size=10, bold=is_tot2,
                       color=LINK_GREEN if isinstance(amt, str) and amt.startswith('=') else INPUT_BLUE)
        c3.alignment = Alignment(horizontal='right', vertical='center')
        if is_tot2: c3.fill = hfill(LIGHT_BLUE)
        ws.cell(r, 4, note).font = Font(name='Calibri', size=8, italic=True, color=DARK_GRAY)
        for col in range(1, 5):
            ws.cell(r, col).border = thin()

# Make the GL verification row special
ws.cell(len(cf_items)+4, 1).font = Font(bold=True, color=ACCENT_GREEN, name='Calibri', size=11)

# ─────────────────────────────────────────────────────────────────────────────
# 12. NOTES
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Notes_Disclosures"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 2, "NOTES TO THE FINANCIAL STATEMENTS — Group 8 | FY 2024", DARK_NAVY, 13)
wcols(ws, {'A':90,'B':20})

notes = [
    ("NOTE 1 — BASIS OF PREPARATION (IAS 1)", MID_BLUE, True),
    ("These financial statements have been prepared in accordance with International Financial Reporting Standards (IFRS) as issued by the IASB. Standards applied include IAS 1 (Revised), IAS 2 (Inventories), IAS 7 (Cash Flows), IAS 16 (PPE), IAS 36 (Impairment), IAS 37 (Provisions), IAS 38 (Intangibles), IAS 32 (Financial Instruments — Equity), and IFRS 16 (Leases). Amounts are stated in Pakistani Rupees (PKR).", None, False),
    ("", None, False),
    ("NOTE 2 — CONTINGENT LIABILITY (IAS 37 / T22)", MID_BLUE, True),
    ("A third-party legal claim exists for which the company cannot make a reliable estimate of the outflow. No provision has been recognised. In accordance with IAS 37.86, this is disclosed as a contingent liability. The claim is assessed as possible but not probable. Legal counsel is reviewing the matter.", None, False),
    ("", None, False),
    ("NOTE 3 — IMPAIRMENT OF MACHINERY (IAS 36 / T23–T24)", MID_BLUE, True),
    ("An impairment indicator (technological obsolescence) was identified. An impairment test compared the Carrying Amount of PKR 19,666,667 (Cost PKR 20,000,000 less Depreciation PKR 333,333) against Recoverable Amount PKR 18,000,000 (higher of Fair Value Less Costs of Disposal and Value in Use). Impairment loss of PKR 2,000,000 was recognised in profit or loss.", None, False),
    ("", None, False),
    ("NOTE 4 — LEASE (IFRS 16 / T18–T20)", MID_BLUE, True),
    ("The company entered a 5-year lease with annual payments of PKR 450,000. Under IFRS 16, a right-of-use asset and lease liability were recognised at PV = PKR 1,597,065, calculated as =PV(13%,5,-450,000). Interest at 13% is accreted annually. Year 1 interest = PKR 207,619. Principal repayment = PKR 242,381.", None, False),
    ("", None, False),
    ("NOTE 5 — INTANGIBLE ASSETS (IAS 38 / T16–T17 / T25)", MID_BLUE, True),
    ("Research expenditure of PKR 75,000 was expensed as incurred (IAS 38.54 — does not qualify for capitalisation). Development costs of PKR 220,000 met all six IAS 38.57 criteria and were capitalised as an intangible asset. Amortisation is over 10 years on a straight-line basis (annual PKR 22,000; monthly PKR 1,833).", None, False),
    ("", None, False),
    ("NOTE 6 — EQUITY TRANSACTIONS (IAS 1 / IAS 32)", MID_BLUE, True),
    ("Share issuance: 550,000 shares @ PKR 28. Share Capital (par PKR 10) = PKR 5,500,000. Share Premium = PKR 9,900,000. Bonus issue of PKR 275,000 capitalised from Retained Earnings. Private Placement: PKR 1,650,000 (Share Capital PKR 500,000 + Share Premium PKR 1,150,000). Treasury shares of PKR 300,000 deducted from equity at cost per IAS 32.33.", None, False),
    ("", None, False),
    ("NOTE 7 — INVENTORY (IAS 2 / T06–T13)", MID_BLUE, True),
    ("Inventories are measured using the FIFO cost formula. Cost includes purchase price and directly attributable costs. Closing inventory is measured at the lower of cost and NRV. NRV write-down of PKR 1,000 was recognised in profit or loss (T13). Batch 1: 100 units @ PKR 1,000; Batch 2: 60 units @ PKR 1,000.", None, False),
    ("", None, False),
    ("NOTE 8 — PROVISIONS (IAS 37 / T21)", MID_BLUE, True),
    ("A provision of PKR 300,000 was recognised for a probable legal obligation arising from a third-party claim where a reliable estimate was available. The provision represents management's best estimate of the expenditure required to settle the present obligation at the reporting date.", None, False),
]

for i, (text, bg, bold) in enumerate(notes):
    r = i + 2
    ws.row_dimensions[r].height = 48 if not bold and text else 24
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, text)
    c.font = Font(name='Calibri', size=10, bold=bold,
                  color=WHITE if bg else BLACK)
    if bg: c.fill = hfill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ─────────────────────────────────────────────────────────────────────────────
# 13. AI LOG
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["AI_Log"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "AI PROMPT LOG — IFRS Accounting Engine Build Record", DARK_NAVY, 13)
hdr_row(ws, 2, ["Step","Prompt Summary","AI Output Summary","Standards Used"])
wcols(ws, {'A':6,'B':44,'C':54,'D':22})

log_data = [
    (1,"Build 35-transaction IFRS journal with Group 8 variables (expanded Dr/Cr rows)",
     "Generated expanded debit/credit entries T01-T35 with D/C flag column for SUMPRODUCT","Multiple"),
    (2,"Create FIFO inventory engine under IAS 2",
     "Batch costing engine, NRV write-down test, COGS and closing inventory formulas","IAS 2"),
    (3,"Build IFRS 16 lease amortisation table",
     "PV =PV(13%,5,-450000), 5-year interest/principal split schedule","IFRS 16"),
    (4,"IAS 16 depreciation and IAS 36 impairment module",
     "SL depreciation 60yr; impairment test: CA vs Recoverable Amount → PKR 2M loss","IAS 16/36"),
    (5,"IAS 38 research vs development split",
     "Research PKR 75k expensed; Dev PKR 220k capitalised, amortised 10yr SL","IAS 38"),
    (6,"General Ledger via SUMPRODUCT from Journal",
     "Used SUMPRODUCT with D/C flag to auto-post debits and credits to ledger","All"),
    (7,"Link GL to Trial Balance and all Financial Statements",
     "Cross-sheet formulas: GL→TB→P&L→SOFP→SCE→CF all auto-update","IAS 1"),
    (8,"Dashboard KPI cards and financial summaries",
     "Live KPIs: Cash, Net Profit, Total Assets, Equity, Current Ratio, Debt Ratio","IAS 1"),
    (9,"Cash Flow — Indirect Method",
     "Net profit → non-cash add-backs → WC changes → investing → financing","IAS 7"),
    (10,"Accounting equation validation and balance checks",
     "=IF(ABS(Assets-(E+L))<1,'BALANCED','ERROR') on Dashboard and SOFP","IAS 1"),
]

for i, (step, prompt, output, ref) in enumerate(log_data):
    r = i + 3
    ws.row_dimensions[r].height = 22
    for col, val in enumerate([step, prompt, output, ref], 1):
        c = ws.cell(r, col, val)
        c.border = thin()
        c.font = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if i % 2 == 0: c.fill = hfill(LIGHT_GRAY)

# ─────────────────────────────────────────────────────────────────────────────
# 14. INPUT FORM
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Input_Form"]
ws.sheet_view.showGridLines = False
title(ws, 1, 1, 4, "TRANSACTION INPUT FORM — Add New Entries", DARK_NAVY, 13)
ws.merge_cells('A2:D2')
ws.cell(2, 1, "Instructions: Fill in the yellow cells below. Then copy the row and paste into the Journal sheet (row after last transaction). The ledger, TB, and all statements update automatically via SUMPRODUCT formulas.").font = Font(name='Calibri', size=9, italic=True, color=DARK_GRAY)
ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[2].height = 36
wcols(ws, {'A':30,'B':34,'C':20,'D':28})

fields = [
    ("Date (YYYY-MM-DD):",           "2024-01-01",   "Format: 2024-MM-DD"),
    ("Transaction ID:",              "T36",           "Unique ID e.g. T36, T37"),
    ("Description:",                 "Enter here",    "Brief description of transaction"),
    ("Account:",                     "Cash",          "Debit or Credit account name — must match GL"),
    ("Dr / Cr:",                     "Dr",            "Type 'Dr' for debit or 'Cr' for credit"),
    ("Amount (PKR):",                0,               "Positive number only"),
    ("IFRS Reference:",              "IAS 1",         "e.g. IAS 1, IAS 2, IFRS 16"),
    ("Category:",                    "Operating",     "Operating / Investing / Financing / Equity / OCI"),
    ("D/C Flag (1=Dr, -1=Cr, 0=Memo):", 1,           "1 for Dr, -1 for Cr, 0 for memo entries"),
]

for i, (label, default, hint) in enumerate(fields):
    r = i + 4
    c1 = ws.cell(r, 1, label); c1.font = Font(bold=True, name='Calibri', size=10); c1.fill = hfill(LIGHT_BLUE)
    c2 = ws.cell(r, 2, default); c2.font = Font(name='Calibri', size=10, color=INPUT_BLUE)
    c2.fill = hfill("FFFDE7"); c2.border = thin()
    if isinstance(default, int): c2.number_format = '#,##0'
    c3 = ws.cell(r, 3, hint); c3.font = Font(name='Calibri', size=9, italic=True, color=DARK_GRAY)

# ─────────────────────────────────────────────────────────────────────────────
# 15. DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws = WS["Dashboard"]
ws.sheet_view.showGridLines = False
ws.sheet_view.showRowColHeaders = False

# Set all column widths
for col in range(1, 15):
    ws.column_dimensions[get_column_letter(col)].width = 13

# Title
ws.merge_cells('A1:N3')
c = ws.cell(1, 1, "GROUP 8  |  IFRS FINANCIAL ACCOUNTING ENGINE  |  FY 2024")
c.font = Font(bold=True, size=20, color=WHITE, name='Calibri')
c.fill = hfill(DARK_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 48; ws.row_dimensions[2].height = 20; ws.row_dimensions[3].height = 8

# Subtitle
ws.merge_cells('A4:N4')
c = ws.cell(4, 1, "Live Dashboard  ●  IAS 1  ●  IAS 2  ●  IAS 7  ●  IAS 16  ●  IAS 36  ●  IAS 38  ●  IFRS 16  ●  All values auto-update from Journal")
c.font = Font(size=9, color=WHITE, name='Calibri', italic=True)
c.fill = hfill(MID_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[4].height = 18

# Accounting equation check — prominent
ws.merge_cells('A5:N5')
eq_formula = f'=IF(ABS(SOFP!C{26}-(SOFP!C{33}+SOFP!C{48}))<1,"✔  ACCOUNTING EQUATION BALANCED  |  Assets = Equity + Liabilities","✘  EQUATION ERROR — Check Journal Entries")'
c = ws.cell(5, 1, eq_formula)
c.font = Font(bold=True, size=11, color=WHITE, name='Calibri')
c.fill = hfill(ACCENT_GREEN)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[5].height = 24

# KPI Section header
ws.merge_cells('A6:N6')
c = ws.cell(6, 1, "KEY PERFORMANCE INDICATORS")
c.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
c.fill = hfill(ACCENT_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[6].height = 20

# KPI Card helper
def kpi(ws, tr, vr, sc, ec, label, formula, fmt, color, icon):
    ws.merge_cells(start_row=tr, start_column=sc, end_row=tr, end_column=ec)
    ws.merge_cells(start_row=vr, start_column=sc, end_row=vr, end_column=ec)
    ws.merge_cells(start_row=vr+1, start_column=sc, end_row=vr+1, end_column=ec)
    ct = ws.cell(tr, sc, f"{icon}  {label}")
    ct.font = Font(bold=True, size=9, color=WHITE, name='Calibri')
    ct.fill = hfill(color); ct.alignment = Alignment(horizontal='center', vertical='center')
    cv = ws.cell(vr, sc, formula)
    cv.font = Font(bold=True, size=15, color=WHITE, name='Calibri')
    cv.fill = hfill(color); cv.number_format = fmt
    cv.alignment = Alignment(horizontal='center', vertical='center')
    cu = ws.cell(vr+1, sc, "PKR" if 'x' not in fmt and '%' not in fmt else ("x" if 'x' in fmt else "%"))
    cu.font = Font(size=8, color=WHITE, name='Calibri', italic=True)
    cu.fill = hfill(color); cu.alignment = Alignment(horizontal='center', vertical='center')

ws.row_dimensions[7].height = 22; ws.row_dimensions[8].height = 34; ws.row_dimensions[9].height = 16

kpi_data = [
    (1, 2,  "Cash Balance",    f"=General_Ledger!F{gl_r['Cash']}",  '#,##0',  ACCENT_GREEN, "💵"),
    (3, 4,  "Net Revenue",     "=SOFP!C26",                          '#,##0',  ACCENT_BLUE,  "📈"),
    (5, 6,  "Net Profit",      f"=Profit_Loss!C{NP_ROW}",           '#,##0',  MID_BLUE,     "💰"),
    (7, 8,  "Total Assets",    "=SOFP!C26",                          '#,##0',  DARK_NAVY,    "🏦"),
    (9,10,  "Total Equity",    "=SOFP!C33",                          '#,##0',  ACCENT_GOLD,  "📊"),
    (11,12, "Current Ratio",   "=IFERROR(SOFP!C24/SOFP!C46,0)",     '0.00x',  ACCENT_RED,   "⚖️"),
    (13,14, "Debt-to-Assets",  "=IFERROR((SOFP!C39+SOFP!C46)/SOFP!C26,0)", '0.0%', DARK_GRAY, "📉"),
]

# Fix: kpi_data references SOFP rows — let me use correct ones
# Total Assets = row 26 → C26 in SOFP (i=22 → r=26)
# Total Equity = row 33 (i=29 → r=33)
# Total Current Assets = C24
# Total Current Liabilities = C46
# Total Non-Current Liabilities = C39

for sc, ec, label, formula, fmt, color, icon in kpi_data:
    kpi(ws, 7, 8, sc, ec, label, formula, fmt, color, icon)

ws.row_dimensions[10].height = 10

# Financial summary section header
ws.merge_cells('A11:N11')
c = ws.cell(11, 1, "FINANCIAL STATEMENTS SUMMARY")
c.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
c.fill = hfill(ACCENT_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[11].height = 20

# P&L Summary (left)
ws.merge_cells('A12:G12')
c = ws.cell(12, 1, "PROFIT & LOSS")
c.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
c.fill = hfill(MID_BLUE); c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[12].height = 20

pl_summary = [
    ("Gross Revenue",          "=Profit_Loss!C5+Profit_Loss!C6"),
    ("Net Revenue",            "=Profit_Loss!C7"),
    ("Total COGS",             "=Profit_Loss!C11"),
    ("Gross Profit",           "=Profit_Loss!C12"),
    ("Total Opex",             "=Profit_Loss!C26"),
    ("EBIT",                   "=Profit_Loss!C28"),
    ("Net Profit",             f"=Profit_Loss!C{NP_ROW}"),
    ("Total Comprehensive Inc","=Profit_Loss!C{last}"),
]

# fix last row
for i2, (desc, *_) in enumerate(pl_items):
    if desc == "TOTAL COMPREHENSIVE INCOME":
        TCI_ROW = i2 + 4
        break

for i, (label, formula) in enumerate(pl_summary):
    formula = formula.replace("{last}", str(TCI_ROW))
    r = i + 13
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    c1 = ws.cell(r, 1, label)
    c1.font = Font(name='Calibri', size=10, bold=(label in ("Gross Profit","Net Profit","Total Comprehensive Inc")))
    c1.alignment = Alignment(horizontal='left', vertical='center')
    c1.border = thin()
    if i % 2 == 0: c1.fill = hfill(LIGHT_GRAY)
    c2 = ws.cell(r, 5, formula)
    c2.number_format = '#,##0'; c2.font = Font(name='Calibri', size=10, color=LINK_GREEN)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin()
    if i % 2 == 0: c2.fill = hfill(LIGHT_GRAY)
    ws.row_dimensions[r].height = 18

# SOFP Summary (right)
ws.merge_cells('H12:N12')
c = ws.cell(12, 8, "FINANCIAL POSITION")
c.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
c.fill = hfill(MID_BLUE); c.alignment = Alignment(horizontal='center', vertical='center')

sofp_summary = [
    ("Total Non-Current Assets", "=SOFP!C17"),
    ("Total Current Assets",     "=SOFP!C24"),
    ("TOTAL ASSETS",             "=SOFP!C26"),
    ("Total Equity",             "=SOFP!C33"),
    ("Non-Current Liabilities",  "=SOFP!C39"),
    ("Current Liabilities",      "=SOFP!C46"),
    ("TOTAL LIABILITIES",        "=SOFP!C48"),
    ("Cash Balance",             f"=General_Ledger!F{gl_r['Cash']}"),
]

for i, (label, formula) in enumerate(sofp_summary):
    r = i + 13
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
    ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=14)
    c1 = ws.cell(r, 8, label)
    c1.font = Font(name='Calibri', size=10, bold=("TOTAL" in label))
    c1.alignment = Alignment(horizontal='left', vertical='center')
    c1.border = thin()
    if i % 2 == 0: c1.fill = hfill(LIGHT_GRAY)
    c2 = ws.cell(r, 11, formula)
    c2.number_format = '#,##0'; c2.font = Font(name='Calibri', size=10, color=LINK_GREEN)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.border = thin()
    if i % 2 == 0: c2.fill = hfill(LIGHT_GRAY)
    ws.row_dimensions[r].height = 18

# Nav bar
ws.row_dimensions[22].height = 8
ws.merge_cells('A23:N23')
c = ws.cell(23, 1, "NAVIGATION — Click a sheet tab below | All 15 sheets auto-linked via Journal → GL → All Statements")
c.font = Font(bold=True, size=9, color=WHITE, name='Calibri', italic=True)
c.fill = hfill(DARK_GRAY); c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[23].height = 20

nav = [
    ("📋 Journal",      ACCENT_RED),
    ("📒 Ledger",       ACCENT_GREEN),
    ("⚖️ Trial Bal",    ACCENT_GOLD),
    ("📦 Inventory",    ACCENT_BLUE),
    ("🏗️ PPE",          MID_BLUE),
    ("📑 Lease",        "16A085"),
    ("🏛️ Equity",       DARK_NAVY),
    ("💰 P&L",          ACCENT_GREEN),
    ("🏦 SOFP",         ACCENT_BLUE),
    ("💸 Cash Flow",    ACCENT_GOLD),
    ("📝 Notes",        DARK_GRAY),
    ("🤖 AI Log",       DARK_NAVY),
    ("🔧 Variables",    MID_BLUE),
    ("📥 Input Form",   ACCENT_RED),
]
for i, (label, color) in enumerate(nav):
    ws.merge_cells(start_row=24, start_column=i+1, end_row=25, end_column=i+1)
    c = ws.cell(24, i+1, label)
    c.font = Font(bold=True, size=8, color=WHITE, name='Calibri')
    c.fill = hfill(color); c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[24].height = 28; ws.row_dimensions[25].height = 8

# Tab colors
tab_cols = {
    "Dashboard": DARK_NAVY, "Variable_Matrix": MID_BLUE, "Journal": ACCENT_RED,
    "General_Ledger": ACCENT_GREEN, "Trial_Balance": ACCENT_GOLD,
    "Inventory_Module": ACCENT_BLUE, "PPE_Module": "8E44AD",
    "Lease_Module": "16A085", "Equity_Module": "D35400",
    "Profit_Loss": ACCENT_BLUE, "SOFP": ACCENT_GREEN,
    "Cash_Flow": ACCENT_GOLD, "Notes_Disclosures": DARK_GRAY,
    "AI_Log": DARK_NAVY, "Input_Form": ACCENT_RED,
}
for name, color in tab_cols.items():
    WS[name].sheet_properties.tabColor = color

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
out = "Group8_IFRS_Accounting_Engine.xlsx"
wb.save(out)
print("Saved:", out)
print(f"NP_ROW={NP_ROW}, TCI_ROW={TCI_ROW}")